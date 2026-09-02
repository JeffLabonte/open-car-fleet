import json
import tempfile
import uuid
from datetime import timedelta
from io import BytesIO
from io import StringIO
from pathlib import Path
from typing import cast
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.middleware import AuthenticationMiddleware
from django.core.management import call_command
from django.core.management.base import CommandError
from django.contrib.sessions.middleware import SessionMiddleware
from django.db import models
from django.http import HttpRequest, HttpResponse
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from shop import views
from shop.exporters import export_garage_to_excel
from shop.forms import CarCreateForm, CarUpdateForm, GarageCreateForm, ReportForm, WorkJobForm
from shop.importers import CSVImporter, ImportContext
from shop.middleware import HankoAuthenticationMiddleware
from shop.models.car import Car, CarPart, CarPartStatusHistory
from shop.models.garage import Garage, GarageInvitation, GarageMembership, KnownShop, KnownShopProof
from shop.models.job import WorkJob
from shop.models.report import Report, ReportAttachment
from shop.models.user import ShopUser


class HankoAuthenticationIntegrationTests(TestCase):
    def test_logged_off_users_are_redirected_to_login_from_app_routes(self):
        response = self.client.get(reverse('shop-index'))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response['Location'].startswith(f"{reverse('shop-login')}?next="))

        login_response = self.client.get(reverse('shop-login'))
        self.assertEqual(login_response.status_code, 200)

    def test_hanko_callback_and_middleware_rehydrate_django_user(self):
        payload = {
            "user": {
                "id": "hanko-user-123",
                "email": "driver@example.com",
                "name": "Test Driver",
                "display_name": "Test Driver",
                "avatar_url": "https://example.com/avatar.png",
                "provider": "hanko",
            }
        }

        response = self.client.post(
            reverse('shop-hanko-callback'),
            data=json.dumps(payload),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])

        user = ShopUser.objects.get(email='driver@example.com')
        self.assertEqual(user.hanko_id, 'hanko-user-123')
        self.assertEqual(user.display_name, 'Test Driver')
        self.assertTrue(user.is_active)
        self.assertFalse(user.garages.exists())

        self.client.logout()
        self.client.session['hanko_session_token'] = 'session-token-123'
        self.client.session.save()

        class FakeResponse:
            def raise_for_status(self) -> None:
                return None

            def json(self) -> dict[str, str]:
                return {
                    'id': 'hanko-user-123',
                    'email': 'driver@example.com',
                    'name': 'Test Driver',
                    'display_name': 'Test Driver',
                    'avatar_url': 'https://example.com/avatar.png',
                    'provider': 'hanko',
                }

        with patch('shop.middleware.requests.get', return_value=FakeResponse()):
            self.client.get(reverse('shop-index'))

        factory = RequestFactory()
        request = factory.get(reverse('shop-index'))

        def next_response(_request: HttpRequest) -> HttpResponse:
            return HttpResponse()

        SessionMiddleware(next_response).process_request(request)
        request.session['hanko_session_token'] = 'session-token-123'
        request.session.save()
        AuthenticationMiddleware(next_response).process_request(request)

        with patch('shop.middleware.requests.get', return_value=FakeResponse()):
            HankoAuthenticationMiddleware(next_response).process_request(request)

        self.assertTrue(request.user.is_authenticated)
        authenticated_user = cast(ShopUser, request.user)
        self.assertEqual(authenticated_user.email, 'driver@example.com')
        self.assertEqual(request.session.get('hanko_user_id'), 'hanko-user-123')
        self.assertEqual(request.session.get('hanko_email'), 'driver@example.com')

        refreshed_user = get_user_model().objects.get(pk=user.pk)
        self.assertEqual(refreshed_user.email, 'driver@example.com')

    def test_logout_clears_hanko_session_and_redirects(self):
        user = ShopUser.objects.create_user(username='logout-user', email='logout@example.com', password='pass1234')
        self.client.force_login(user)
        self.client.session['hanko_session_token'] = 'session-token-123'
        self.client.session['hanko_user_id'] = 'hanko-user-123'
        self.client.session.save()

        response = self.client.post(reverse('shop-logout'))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('shop-login'))
        self.assertNotIn('hanko_session_token', self.client.session)
        self.assertNotIn('hanko_user_id', self.client.session)

    def test_logout_marker_is_one_time_in_login_context(self):
        user = ShopUser.objects.create_user(username='logout-user-2', email='logout2@example.com', password='pass1234')
        self.client.force_login(user)

        self.client.post(reverse('shop-logout'))

        first_login_page = self.client.get(reverse('shop-login'))
        second_login_page = self.client.get(reverse('shop-login'))

        self.assertEqual(first_login_page.context['logged_out'], True)
        self.assertEqual(second_login_page.context['logged_out'], False)

    @override_settings(
        ALLOWED_HOSTS=['xps-server.kanyu-bluegill.ts.net'],
        CSRF_TRUSTED_ORIGINS=['https://xps-server.kanyu-bluegill.ts.net'],
    )
    def test_logout_with_trusted_origin_succeeds(self):
        from django.test import Client
        from django.middleware.csrf import get_token
        client = Client(enforce_csrf_checks=True)
        user = ShopUser.objects.create_user(username='csrf-user', email='csrf@example.com', password='pass1234')
        client.force_login(user)

        # login_view doesn't render a {% csrf_token %} tag, so no CSRF cookie
        # is ever set by the response. Generate a token directly and seed it
        # on the client so both the cookie and the submitted form field agree.
        csrf_token = get_token(RequestFactory().get('/'))
        client.cookies['csrftoken'] = csrf_token

        # Untrusted origin fails CSRF check with 403
        untrusted_resp = client.post(
            reverse('shop-logout'),
            {'csrfmiddlewaretoken': csrf_token},
            HTTP_ORIGIN='https://untrusted-domain.com',
            HTTP_HOST='xps-server.kanyu-bluegill.ts.net',
        )
        self.assertEqual(untrusted_resp.status_code, 403)

        # Trusted origin succeeds with 302
        trusted_resp = client.post(
            reverse('shop-logout'),
            {'csrfmiddlewaretoken': csrf_token},
            HTTP_ORIGIN='https://xps-server.kanyu-bluegill.ts.net',
            HTTP_HOST='xps-server.kanyu-bluegill.ts.net',
        )
        self.assertEqual(trusted_resp.status_code, 302)

    def test_theme_preference_sets_cookie_and_redirects(self):
        response = self.client.get(
            reverse('shop-theme', kwargs={'theme': 'dark'}),
            {'next': reverse('shop-login')},
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('shop-login'))
        self.assertEqual(response.cookies['theme'].value, 'dark')
        self.assertEqual(response.cookies['theme']['max-age'], '31536000')

    def test_login_page_renders_current_theme_attribute(self):
        response = self.client.get(reverse('shop-login'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-theme="light"')

    def test_login_page_includes_mobile_viewport_and_theme_toggle(self):
        response = self.client.get(reverse('shop-login'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<meta name="viewport"')
        self.assertContains(response, 'Light mode')
        self.assertContains(response, 'Dark mode')
        self.assertContains(response, 'beta')
        self.assertContains(response, 'theme-beta')

    @override_settings(HANKO_API_URL='https://hanko.example.com')
    def test_car_list_uses_hanko_session_token_to_authenticate(self):
        factory = RequestFactory()
        request = factory.get(reverse('shop-car-list'))

        def next_response(_request: HttpRequest) -> HttpResponse:
            return HttpResponse()

        SessionMiddleware(next_response).process_request(request)
        request.session['hanko_session_token'] = 'session-token-123'
        request.session.save()
        AuthenticationMiddleware(next_response).process_request(request)

        class FakeResponse:
            def raise_for_status(self) -> None:
                return None

            def json(self) -> dict[str, str]:
                return {
                    'id': 'hanko-user-123',
                    'email': 'driver@example.com',
                    'name': 'Test Driver',
                    'display_name': 'Test Driver',
                    'avatar_url': 'https://example.com/avatar.png',
                    'provider': 'hanko',
                }

        with patch('shop.middleware.requests.get', return_value=FakeResponse()):
            response = views.car_list(request)

        self.assertEqual(response.status_code, 200)
        self.assertIn(b'Cars', response.content)


class CarPartStatusTrackingTests(TestCase):
    def setUp(self) -> None:
        self.garage = Garage.objects.create(name='North Garage')
        self.car = Car.objects.create(
            garage=self.garage,
            make='Toyota',
            model='Corolla',
            colour='Blue',
            year=2022,
            vin='1HGBH41JXMN109186',
            license_plate='ABC123',
        )

    def test_part_status_changes_are_recorded_with_timestamps(self):
        part = CarPart.objects.create(
            car=self.car,
            name='Brake pads',
            status=CarPart.STATUS_NEW,
            notes='Initial issue spotted on inspection.',
        )

        self.assertEqual(part.status, CarPart.STATUS_NEW)
        self.assertEqual(part.status_history.count(), 1)

        part.update_status(CarPart.STATUS_ORDERED, note='Ordered replacement set from supplier.')
        part.refresh_from_db()

        self.assertEqual(part.status, CarPart.STATUS_ORDERED)
        self.assertEqual(part.status_history.count(), 2)

        first_event = part.status_history.order_by('changed_at').first()
        second_event = part.status_history.order_by('changed_at').last()

        self.assertEqual(first_event.previous_status, '')
        self.assertEqual(first_event.new_status, CarPart.STATUS_NEW)
        self.assertIsNotNone(first_event.changed_at)

        self.assertEqual(second_event.previous_status, CarPart.STATUS_NEW)
        self.assertEqual(second_event.new_status, CarPart.STATUS_ORDERED)
        self.assertEqual(second_event.note, 'Ordered replacement set from supplier.')
        self.assertIsNotNone(second_event.changed_at)


class GarageSharingTests(TestCase):
    def setUp(self) -> None:
        self.owner = ShopUser.objects.create_user(
            username='owner',
            email='owner@example.com',
            password='pass1234',
        )
        self.member = ShopUser.objects.create_user(
            username='member',
            email='member@example.com',
            password='pass1234',
        )
        self.stranger = ShopUser.objects.create_user(
            username='stranger',
            email='stranger@example.com',
            password='pass1234',
        )
        self.garage = Garage.objects.create(name='Alpha Garage', created_by=self.owner)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.owner,
            role=GarageMembership.ROLE_OWNER,
        )


    def test_create_garage_adds_owner_membership(self):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('shop-garage-create'),
            data={'name': 'Second Garage', 'description': 'Family vehicles'},
        )

        self.assertEqual(response.status_code, 302)
        created = Garage.objects.get(name='Second Garage')
        self.assertTrue(
            GarageMembership.objects.filter(
                garage=created,
                user=self.owner,
                role=GarageMembership.ROLE_OWNER,
            ).exists()
        )

    @patch('shop.models.garage.send_mail', return_value=1)
    def test_share_garage_creates_pending_invitation(self, _mock_send_mail: object):
        self.client.force_login(self.owner)

        response = self.client.post(
            reverse('shop-garage-share', args=[self.garage.pk]),
            data={
                'invited_email': 'member@example.com',
                'message': 'Join this garage.',
                'expires_in_days': 14,
            },
        )

        self.assertEqual(response.status_code, 302)
        invitation = GarageInvitation.objects.get(garage=self.garage, invited_email='member@example.com')
        self.assertEqual(invitation.status, GarageInvitation.STATUS_PENDING)
        self.assertEqual(invitation.invited_by, self.owner)
        self.assertIsNotNone(invitation.expires_at)

    @patch('shop.models.garage.send_mail', return_value=1)
    def test_duplicate_pending_invitation_is_not_created(self, _mock_send_mail: object):
        self.client.force_login(self.owner)
        GarageInvitation.objects.create(
            garage=self.garage,
            invited_email='member@example.com',
            invited_by=self.owner,
            status=GarageInvitation.STATUS_PENDING,
            expires_at=timezone.now() + timedelta(days=14),
        )

        response = self.client.post(
            reverse('shop-garage-share', args=[self.garage.pk]),
            data={
                'invited_email': 'member@example.com',
                'message': 'Second invite',
                'expires_in_days': 14,
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            GarageInvitation.objects.filter(
                garage=self.garage,
                invited_email='member@example.com',
                status=GarageInvitation.STATUS_PENDING,
            ).count(),
            1,
        )

    @patch('shop.models.garage.send_mail', return_value=1)
    def test_accept_invitation_adds_membership(self, _mock_send_mail: object):
        invitation = GarageInvitation.objects.create(
            garage=self.garage,
            invited_email='member@example.com',
            invited_by=self.owner,
            status=GarageInvitation.STATUS_PENDING,
            expires_at=timezone.now() + timedelta(days=14),
        )
        self.client.force_login(self.member)

        response = self.client.get(reverse('shop-garage-invitation-accept', args=[invitation.token]))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            GarageMembership.objects.filter(
                garage=self.garage,
                user=self.member,
            ).exists()
        )
        invitation.refresh_from_db()
        self.assertEqual(invitation.status, GarageInvitation.STATUS_ACCEPTED)
        self.assertEqual(invitation.accepted_by, self.member)

    def test_accept_invitation_requires_matching_email(self):
        invitation = GarageInvitation.objects.create(
            garage=self.garage,
            invited_email='member@example.com',
            invited_by=self.owner,
            status=GarageInvitation.STATUS_PENDING,
            expires_at=timezone.now() + timedelta(days=14),
        )
        self.client.force_login(self.stranger)

        response = self.client.get(reverse('shop-garage-invitation-accept', args=[invitation.token]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            GarageMembership.objects.filter(
                garage=self.garage,
                user=self.stranger,
            ).exists()
        )
        invitation.refresh_from_db()
        self.assertEqual(invitation.status, GarageInvitation.STATUS_PENDING)

    def test_non_manager_cannot_share_garage(self):
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.member,
            role=GarageMembership.ROLE_MEMBER,
        )
        self.client.force_login(self.member)

        response = self.client.get(reverse('shop-garage-share', args=[self.garage.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('shop-garage-detail', args=[self.garage.pk]))

    def test_owner_can_delete_car_from_car_list(self):
        self.client.force_login(self.owner)
        car = Car.objects.create(
            garage=self.garage,
            make='Toyota',
            model='Yaris',
            vin='JTDKB20U793512345',
        )

        response = self.client.post(reverse('shop-car-delete', args=[car.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('shop-car-list'))
        self.assertFalse(Car.objects.filter(pk=car.pk).exists())


class KnownShopTests(TestCase):
    def setUp(self) -> None:
        self.user = ShopUser.objects.create_user(
            username='shop-user',
            email='shop-user@example.com',
            password='pass1234',
        )

    def test_user_can_add_shop_and_proof(self):
        self.client.force_login(self.user)

        shop_response = self.client.post(
            reverse('shop-known-shop-create'),
            data={
                'name': 'Northside Auto',
                'email': 'service@northside.example',
                'phone': '555-0100',
                'address': '10 Main Street',
                'notes': 'Recommended by the fleet manager.',
            },
        )

        self.assertEqual(shop_response.status_code, 302)
        shop = KnownShop.objects.get(name='Northside Auto')
        proof_response = self.client.post(
            reverse('shop-known-shop-proof-create', args=[shop.pk]),
            data={
                'title': 'Business registration',
                'content': 'Registration document received.',
                'file': SimpleUploadedFile('registration.pdf', b'%PDF-1.4 proof', content_type='application/pdf'),
            },
        )

        self.assertEqual(proof_response.status_code, 302)
        proof = KnownShopProof.objects.get(shop=shop)
        self.assertEqual(proof.title, 'Business registration')
        self.assertIn('registration', proof.file.name)
        self.assertTrue(proof.file.name.endswith('.pdf'))


class FormEditableFieldsCoverageTests(TestCase):
    def _editable_model_field_names(self, model: type[models.Model], *, exclude: set[str] | None = None) -> set[str]:
        excluded = exclude or set()
        return {
            field.name
            for field in model._meta.fields
            if field.editable and not field.auto_created and field.name not in excluded
        }

    def test_car_create_form_covers_editable_car_fields(self):
        expected = self._editable_model_field_names(Car)
        self.assertSetEqual(set(CarCreateForm.base_fields.keys()), expected)

    def test_car_update_form_covers_editable_car_fields(self):
        expected = self._editable_model_field_names(Car)
        self.assertSetEqual(set(CarUpdateForm.base_fields.keys()), expected)

    def test_garage_create_form_covers_user_editable_garage_fields(self):
        expected = self._editable_model_field_names(Garage, exclude={'created_by'})
        self.assertSetEqual(set(GarageCreateForm.base_fields.keys()), expected)

    def test_workjob_form_covers_user_editable_workjob_fields(self):
        expected = self._editable_model_field_names(WorkJob, exclude={'car'})
        self.assertSetEqual(set(WorkJobForm.base_fields.keys()), expected)

    def test_report_form_covers_user_editable_report_fields(self):
        expected = self._editable_model_field_names(Report, exclude={'car'})
        actual = set(ReportForm.base_fields.keys())
        self.assertTrue(expected.issubset(actual))
        self.assertSetEqual(actual - expected, {'attachments', 'external_links'})


class ColourFieldTests(TestCase):
    def setUp(self) -> None:
        self.user = ShopUser.objects.create_user(
            username='colour-owner',
            email='colour-owner@example.com',
            password='pass1234',
        )
        self.garage = Garage.objects.create(name='Colour Garage', created_by=self.user)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.user,
            role=GarageMembership.ROLE_OWNER,
        )

    def test_car_persists_colour_field(self):
        car = Car.objects.create(
            garage=self.garage,
            make='Toyota',
            model='Yaris',
            colour='Noir',
            vin='JTDKB20U793512346',
        )
        car.refresh_from_db()
        self.assertEqual(car.colour, 'Noir')

    def test_car_create_form_includes_colour_field_with_british_label(self):
        self.assertIn('colour', CarCreateForm.base_fields)
        form = CarCreateForm(user=self.user)
        self.assertEqual(form.fields['colour'].label, 'Colour')

    def test_importer_reads_colour_key_from_record(self):
        importer = CSVImporter()
        result = importer.import_records(
            Car,
            [{
                'make': 'Toyota',
                'model': 'Yaris',
                'colour': 'Blanc',
                'vin': 'JTDKB20U793512347',
            }],
            context=ImportContext(garage=self.garage),
        )

        self.assertFalse(result.has_errors)
        self.assertEqual(result.created_count, 1)
        car = Car.objects.get(vin='JTDKB20U793512347')
        self.assertEqual(car.colour, 'Blanc')

    def test_car_list_renders_colour_label_and_value(self):
        self.client.force_login(self.user)
        Car.objects.create(
            garage=self.garage,
            make='Toyota',
            model='Yaris',
            colour='Noir',
            vin='JTDKB20U793512348',
        )

        response = self.client.get(reverse('shop-car-list'))

        self.assertEqual(response.status_code, 200)
        self.assertIn(b'Colour:', response.content)
        self.assertIn(b'Noir', response.content)

    def test_car_detail_renders_colour_label_and_value(self):
        self.client.force_login(self.user)
        car = Car.objects.create(
            garage=self.garage,
            make='Toyota',
            model='Yaris',
            colour='Noir',
            vin='JTDKB20U793512349',
        )

        response = self.client.get(reverse('shop-car-detail', args=[car.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertIn(b'Colour:', response.content)
        self.assertIn(b'Noir', response.content)


class ReportAttachmentTests(TestCase):
    def setUp(self) -> None:
        self.user = ShopUser.objects.create_user(
            username='report-owner',
            email='report-owner@example.com',
            password='pass1234',
            is_mechanic=True,
        )
        self.garage = Garage.objects.create(name='Report Garage', created_by=self.user)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.user,
            role=GarageMembership.ROLE_OWNER,
        )
        self.car = Car.objects.create(
            garage=self.garage,
            make='Toyota',
            model='Prius',
            vin='JTDKB20U123456789',
        )

    def test_report_create_accepts_uploads_and_external_links(self):
        self.client.force_login(self.user)
        image = SimpleUploadedFile('before.png', b'fake-image', content_type='image/png')
        video = SimpleUploadedFile('clip.mp4', b'fake-video', content_type='video/mp4')

        job_name = f'Brake service {self._testMethodName}'
        response = self.client.post(
            reverse('shop-report-create', args=[self.car.pk]),
            data={
                'mileage': '125000',
                'job_name': job_name,
                'date_done': '2026-08-09',
                'note': 'Replaced brake pads',
                'additional_information': 'Used OEM parts and torqued wheels to spec',
                'external_links': 'https://onedrive.example.com/share/abc\nhttps://drive.google.com/file/d/123/view',
                'attachments': [image, video],
            },
        )

        self.assertEqual(response.status_code, 302)
        report = Report.objects.get(job_name=job_name)
        self.assertEqual(report.attachments.count(), 4)
        self.assertEqual(report.attachments.filter(source_type='upload').count(), 2)
        self.assertEqual(report.attachments.filter(source_type='external').count(), 2)
        self.assertEqual(report.additional_information, 'Used OEM parts and torqued wheels to spec')
        self.assertTrue(report.attachments.filter(source_type='external').exists())
        self.assertTrue(report.attachments.filter(source_type='upload', kind='image').exists())
        self.assertTrue(report.attachments.filter(source_type='upload', kind='video').exists())
        attachment_urls = {attachment.url for attachment in report.attachments.all()}
        self.assertIn('https://onedrive.example.com/share/abc', attachment_urls)
        self.assertIn('https://drive.google.com/file/d/123/view', attachment_urls)

    def test_car_detail_renders_attachment_preview_links(self):
        self.client.force_login(self.user)
        report = Report.objects.create(
            car=self.car,
            job_name='Oil change',
            date_done='2026-08-10',
            note='Completed',
        )
        ReportAttachment.objects.create(
            report=report,
            source_type='external',
            url='https://drive.google.com/file/d/456/view',
            display_name='Service checklist',
            kind='link',
        )

        response = self.client.get(reverse('shop-car-detail', args=[self.car.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Service checklist')
        self.assertContains(response, 'https://drive.google.com/file/d/456/view')


class CSVImporterTests(TestCase):
    def setUp(self) -> None:
        self.user = ShopUser.objects.create_user(
            username='import-owner',
            email='import-owner@example.com',
            password='pass1234',
        )
        self.garage = Garage.objects.create(name='Import Garage', created_by=self.user)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.user,
            role=GarageMembership.ROLE_OWNER,
        )
        self.importer = CSVImporter()

    def test_car_dry_run_requires_target_garage_and_does_not_persist(self):
        result = self.importer.import_records(
            Car,
            [{'make': 'Toyota', 'model': 'Yaris', 'vin': 'JTDKB20U793512345'}],
            context=ImportContext(garage=self.garage),
            dry_run=True,
        )

        self.assertFalse(result.has_errors)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(Car.objects.count(), 0)

    def test_report_import_parses_lists_and_persists(self):
        car = Car.objects.create(garage=self.garage, make='Honda', model='Civic', vin='2HGFG12698H512345')

        result = self.importer.import_records(
            Report,
            [{
                'car': str(car.pk),
                'job_name': 'Brake service',
                'date_done': '2026-08-01',
                'documents': 'invoice.pdf\nchecklist.pdf',
                'photos': 'before.jpg\nafter.jpg',
                'mileage': '12345',
            }],
            context=ImportContext(garage=self.garage),
        )

        self.assertFalse(result.has_errors)
        self.assertEqual(result.created_count, 1)
        report_data = Report.objects.filter(job_name='Brake service').values('documents', 'photos', 'mileage').get()
        self.assertEqual(report_data['documents'], ['invoice.pdf', 'checklist.pdf'])
        self.assertEqual(report_data['photos'], ['before.jpg', 'after.jpg'])
        self.assertEqual(report_data['mileage'], 12345)

    def test_workjob_import_rejects_unknown_car(self):
        result = self.importer.import_records(
            WorkJob,
            [{'car': 'missing-car', 'title': 'Oil change'}],
            context=ImportContext(garage=self.garage),
            dry_run=True,
        )

        self.assertTrue(result.has_errors)
        self.assertIn('Car not found', result.errors[0].message)


class ImportCsvCommandTests(TestCase):
    def setUp(self) -> None:
        self.user = ShopUser.objects.create_user(
            username='command-owner',
            email='command-owner@example.com',
            password='pass1234',
        )
        self.garage = Garage.objects.create(name='Command Garage', created_by=self.user)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.user,
            role=GarageMembership.ROLE_OWNER,
        )

    def test_car_import_command_dry_run_validates_without_persisting(self):
        with tempfile.NamedTemporaryFile('w', suffix='.csv', delete=False) as handle:
            handle.write('make,model,vin\nMazda,3,JM1BK323171512345\n')
            temp_path = handle.name

        output = StringIO()
        try:
            call_command(
                'import_csv',
                'Car',
                temp_path,
                '--garage',
                str(self.garage.pk),
                '--dry-run',
                stdout=output,
            )
        finally:
            Path(temp_path).unlink(missing_ok=True)

        self.assertIn('Dry run complete', output.getvalue())
        self.assertEqual(Car.objects.count(), 0)

    def test_export_garage_command_writes_excel_file(self):
        Car.objects.create(
            garage=self.garage,
            usual_name='Command Export Car',
            make='Mazda',
            model='3',
            vin='JM1BK323171512345',
        )
        output = StringIO()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as handle:
            output_path = Path(handle.name)
        output_path.unlink(missing_ok=True)

        try:
            call_command(
                'export_garage',
                str(self.garage.pk),
                '--output',
                str(output_path),
                stdout=output,
            )
            self.assertTrue(output_path.exists())
            workbook = load_workbook(filename=str(output_path))
            self.assertIn('cars_import', workbook.sheetnames)
            cars_sheet = workbook['cars_import']
            rows = list(cars_sheet.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0][3], 'Mazda')
        finally:
            output_path.unlink(missing_ok=True)

    def test_export_garage_command_errors_for_unknown_garage(self):
        with self.assertRaises(CommandError):
            call_command('export_garage', str(uuid.uuid4()))


class GarageExportServiceTests(TestCase):
    def setUp(self) -> None:
        self.owner = ShopUser.objects.create_user(
            username='export-owner',
            email='export-owner@example.com',
            password='pass1234',
            is_mechanic=True,
        )
        self.garage = Garage.objects.create(name='Primary Garage', created_by=self.owner)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.owner,
            role=GarageMembership.ROLE_OWNER,
        )

        self.other_garage = Garage.objects.create(name='Other Garage', created_by=self.owner)
        GarageMembership.objects.create(
            garage=self.other_garage,
            user=self.owner,
            role=GarageMembership.ROLE_OWNER,
        )

        self.primary_car = Car.objects.create(
            garage=self.garage,
            usual_name='Daily Driver',
            make='Toyota',
            model='Corolla',
            vin='2T1BURHE5JC512345',
        )
        self.other_car = Car.objects.create(
            garage=self.other_garage,
            usual_name='Spare Car',
            make='Honda',
            model='Civic',
            vin='2HGFG12698H512345',
        )

        WorkJob.objects.create(
            car=self.primary_car,
            title='Oil Change',
            assigned_to=self.owner,
            required_items=['Oil', 'Filter'],
            status='pending',
            urgency='soon',
        )
        WorkJob.objects.create(
            car=self.other_car,
            title='Do Not Export',
        )

        Report.objects.create(
            car=self.primary_car,
            mileage=120000,
            job_name='Brake Service',
            assigned_to=self.owner,
            date_done=timezone.now().date(),
            documents=['invoice.pdf'],
            photos=['before.jpg', 'after.jpg'],
        )
        Report.objects.create(
            car=self.other_car,
            job_name='Skip Report',
            date_done=timezone.now().date(),
        )

    def test_export_garage_to_excel_includes_expected_sheets_and_rows(self):
        workbook_file = export_garage_to_excel(self.garage)
        workbook = load_workbook(filename=BytesIO(workbook_file.content))

        self.assertIn('meta', workbook.sheetnames)
        self.assertIn('garage', workbook.sheetnames)
        self.assertIn('memberships', workbook.sheetnames)
        self.assertIn('cars_import', workbook.sheetnames)
        self.assertIn('workjobs_import', workbook.sheetnames)
        self.assertIn('reports_import', workbook.sheetnames)

        cars_rows = list(workbook['cars_import'].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(cars_rows), 1)
        self.assertEqual(cars_rows[0][3], 'Toyota')
        self.assertNotIn('2HGFG12698H512345', [row[7] for row in cars_rows])

        job_rows = list(workbook['workjobs_import'].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(job_rows), 1)
        self.assertEqual(job_rows[0][3], 'Oil Change')
        self.assertEqual(job_rows[0][14], 'Oil\nFilter')

        report_rows = list(workbook['reports_import'].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(report_rows), 1)
        self.assertEqual(report_rows[0][4], 'Brake Service')
        self.assertEqual(report_rows[0][12], 'invoice.pdf')
        self.assertEqual(report_rows[0][13], 'before.jpg\nafter.jpg')


class GarageExportViewTests(TestCase):
    def setUp(self) -> None:
        self.owner = ShopUser.objects.create_user(
            username='export-view-owner',
            email='export-view-owner@example.com',
            password='pass1234',
            is_mechanic=True,
        )
        self.manager = ShopUser.objects.create_user(
            username='export-view-manager',
            email='export-view-manager@example.com',
            password='pass1234',
        )
        self.member = ShopUser.objects.create_user(
            username='export-view-member',
            email='export-view-member@example.com',
            password='pass1234',
        )

        self.garage = Garage.objects.create(name='Export View Garage', created_by=self.owner)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.owner,
            role=GarageMembership.ROLE_OWNER,
        )
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.manager,
            role=GarageMembership.ROLE_MANAGER,
        )
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.member,
            role=GarageMembership.ROLE_MEMBER,
        )

        self.primary_car = Car.objects.create(
            garage=self.garage,
            usual_name='Exportable',
            make='Ford',
            model='Focus',
            vin='1FAHP3F28CL512345',
        )

        self.other_garage = Garage.objects.create(name='External Garage', created_by=self.owner)
        GarageMembership.objects.create(
            garage=self.other_garage,
            user=self.owner,
            role=GarageMembership.ROLE_OWNER,
        )
        Car.objects.create(
            garage=self.other_garage,
            usual_name='Hidden',
            make='Tesla',
            model='Model 3',
            vin='5YJ3E1EA7LF512345',
        )

    def test_unauthenticated_user_is_redirected_from_garage_export(self):
        response = self.client.get(reverse('shop-garage-export', args=[self.garage.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response['Location'].startswith(f"{reverse('shop-login')}?next="))

    def test_non_manager_is_redirected_from_garage_export(self):
        self.client.force_login(self.member)

        response = self.client.get(reverse('shop-garage-export', args=[self.garage.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('shop-garage-detail', args=[self.garage.pk]))

    def test_manager_can_download_garage_export_workbook(self):
        self.client.force_login(self.manager)

        response = self.client.get(reverse('shop-garage-export', args=[self.garage.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment; filename="garage_export-view-garage_', response['Content-Disposition'])

        workbook = load_workbook(filename=BytesIO(response.content))
        cars_rows = list(workbook['cars_import'].iter_rows(min_row=2, values_only=True))
        vins = [row[7] for row in cars_rows]
        self.assertIn('1FAHP3F28CL512345', vins)
        self.assertNotIn('5YJ3E1EA7LF512345', vins)


class GarageImportViewTests(TestCase):
    def setUp(self) -> None:
        self.owner = ShopUser.objects.create_user(
            username='garage-owner',
            email='garage-owner@example.com',
            password='pass1234',
        )
        self.member = ShopUser.objects.create_user(
            username='garage-member',
            email='garage-member@example.com',
            password='pass1234',
        )
        self.garage = Garage.objects.create(name='Upload Garage', created_by=self.owner)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.owner,
            role=GarageMembership.ROLE_OWNER,
        )
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.member,
            role=GarageMembership.ROLE_MEMBER,
        )

    def test_non_manager_is_redirected_from_garage_import(self):
        self.client.force_login(self.member)

        response = self.client.get(reverse('shop-garage-import', args=[self.garage.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('shop-garage-detail', args=[self.garage.pk]))

    def test_manager_can_dry_run_car_import_from_upload(self):
        self.client.force_login(self.owner)
        upload = SimpleUploadedFile(
            'cars.csv',
            b'make,model,vin\nSubaru,Outback,4S4BSENC0J3351234',
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('shop-garage-import', args=[self.garage.pk]),
            data={
                'import_file': upload,
                'dry_run': 'on',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        messages = list(response.context['messages'])
        self.assertTrue(any('Dry run complete' in str(message) for message in messages))
        self.assertEqual(Car.objects.count(), 0)

    def test_manager_can_import_cars_into_selected_garage(self):
        self.client.force_login(self.owner)
        upload = SimpleUploadedFile(
            'cars.csv',
            b'make,model,vin\nFord,Focus,1FAHP3F28CL512345',
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('shop-garage-import', args=[self.garage.pk]),
            data={
                'import_file': upload,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        car = Car.objects.get(vin='1FAHP3F28CL512345')
        self.assertEqual(car.garage, self.garage)


class CarImportViewTests(TestCase):
    def setUp(self) -> None:
        self.owner = ShopUser.objects.create_user(
            username='car-owner',
            email='car-owner@example.com',
            password='pass1234',
        )
        self.member = ShopUser.objects.create_user(
            username='car-member',
            email='car-member@example.com',
            password='pass1234',
        )
        self.garage = Garage.objects.create(name='Car Import Garage', created_by=self.owner)
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.owner,
            role=GarageMembership.ROLE_OWNER,
        )
        GarageMembership.objects.create(
            garage=self.garage,
            user=self.member,
            role=GarageMembership.ROLE_MEMBER,
        )
        self.car = Car.objects.create(
            garage=self.garage,
            usual_name='Daily Driver',
            make='Toyota',
            model='Corolla',
            vin='2T1BURHE5JC512345',
        )

    def test_non_manager_is_redirected_from_car_import(self):
        self.client.force_login(self.member)

        response = self.client.get(reverse('shop-car-import', args=[self.car.pk]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('shop-car-detail', args=[self.car.pk]))

    def test_manager_can_dry_run_workjob_import_for_selected_car(self):
        self.client.force_login(self.owner)
        upload = SimpleUploadedFile(
            'workjobs.csv',
            b'title,planned_date\nOil change,2026-08-02',
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('shop-car-import', args=[self.car.pk]),
            data={
                'import_type': 'workjob',
                'import_file': upload,
                'dry_run': 'on',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        messages = list(response.context['messages'])
        self.assertTrue(any('Dry run complete' in str(message) for message in messages))
        self.assertEqual(WorkJob.objects.count(), 0)

    def test_manager_can_import_report_for_selected_car_without_car_field(self):
        self.client.force_login(self.owner)
        upload = SimpleUploadedFile(
            'reports.csv',
            b'job_name,date_done,note\nBrake service,2026-08-03,Pads replaced',
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('shop-car-import', args=[self.car.pk]),
            data={
                'import_type': 'report',
                'import_file': upload,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        report = Report.objects.get(job_name='Brake service')
        self.assertEqual(report.car, self.car)
