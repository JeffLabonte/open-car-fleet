import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from django.utils import timezone

from openpyxl import Workbook

from shop.models.car import Car
from shop.models.garage import Garage
from shop.models.job import WorkJob
from shop.models.report import Report


@dataclass(slots=True)
class GarageExportWorkbook:
    filename: str
    content: bytes


def export_garage_to_excel(garage: Garage) -> GarageExportWorkbook:
    exported_at = timezone.now()
    cars = list(Car.objects.filter(garage=garage).order_by("created_at"))
    jobs = list(
        WorkJob.objects.filter(car__garage=garage)
        .select_related("car", "assigned_to", "assigned_shop")
        .order_by("id")
    )
    reports = list(
        Report.objects.filter(car__garage=garage)
        .select_related("car", "assigned_to", "assigned_shop")
        .order_by("id")
    )
    memberships = list(garage.memberships.select_related("user").order_by("joined_at", "id"))

    workbook = Workbook()
    meta_sheet = workbook.active
    meta_sheet.title = "meta"
    _append_rows(
        meta_sheet,
        ["key", "value"],
        [
            ["export_version", "1"],
            ["exported_at_utc", exported_at.isoformat()],
            ["garage_id", str(garage.pk)],
            ["garage_name", garage.name],
            [
                "row_counts_json",
                json.dumps(
                    {
                        "memberships": len(memberships),
                        "cars_import": len(cars),
                        "workjobs_import": len(jobs),
                        "reports_import": len(reports),
                    }
                ),
            ],
        ],
    )

    garage_sheet = workbook.create_sheet("garage")
    _append_rows(
        garage_sheet,
        [
            "garage_id",
            "name",
            "description",
            "created_by_user_id",
            "created_at",
            "updated_at",
        ],
        [
            [
                str(garage.pk),
                garage.name,
                garage.description,
                _string_or_empty(garage.created_by_id),
                _format_datetime(garage.created_at),
                _format_datetime(garage.updated_at),
            ]
        ],
    )

    memberships_sheet = workbook.create_sheet("memberships")
    _append_rows(
        memberships_sheet,
        [
            "garage_id",
            "user_id",
            "user_email",
            "username",
            "display_name",
            "role",
            "joined_at",
        ],
        [
            [
                str(garage.pk),
                _string_or_empty(membership.user_id),
                membership.user.email,
                membership.user.username,
                membership.user.display_name,
                membership.role,
                _format_datetime(membership.joined_at),
            ]
            for membership in memberships
        ],
    )

    cars_sheet = workbook.create_sheet("cars_import")
    _append_rows(
        cars_sheet,
        [
            "car_id",
            "garage_id",
            "usual_name",
            "make",
            "model",
            "color",
            "year",
            "mileage",
            "vin",
            "license_plate",
            "created_at",
            "updated_at",
        ],
        [
            [
                str(car.pk),
                str(garage.pk),
                car.usual_name,
                car.make,
                car.model,
                car.color,
                _string_or_empty(car.year),
                _string_or_empty(car.mileage),
                car.vin or "",
                car.license_plate,
                _format_datetime(car.created_at),
                _format_datetime(car.updated_at),
            ]
            for car in cars
        ],
    )

    jobs_sheet = workbook.create_sheet("workjobs_import")
    _append_rows(
        jobs_sheet,
        [
            "workjob_id",
            "car_id",
            "car_ref_for_import",
            "title",
            "maintenance_type",
            "assigned_to_user_id",
            "assigned_to_email",
            "assigned_to_username",
            "assigned_shop_id",
            "assigned_shop_name",
            "assigned_shop_email",
            "planned_date",
            "is_done",
            "done_date",
            "required_items_text",
            "urgency",
            "status",
            "notes",
            "created_at",
        ],
        [
            [
                str(job.pk),
                str(job.car_id),
                _car_reference(job.car),
                job.title,
                job.maintenance_type,
                _string_or_empty(job.assigned_to_id),
                job.assigned_to.email if job.assigned_to else "",
                job.assigned_to.username if job.assigned_to else "",
                _string_or_empty(job.assigned_shop_id),
                job.assigned_shop.name if job.assigned_shop else "",
                job.assigned_shop.email if job.assigned_shop else "",
                _format_date(job.planned_date),
                bool(job.is_done),
                _format_date(job.done_date),
                _join_list(job.required_items),
                job.urgency,
                job.status,
                job.notes,
                _format_datetime(job.created_at),
            ]
            for job in jobs
        ],
    )

    reports_sheet = workbook.create_sheet("reports_import")
    _append_rows(
        reports_sheet,
        [
            "report_id",
            "car_id",
            "car_ref_for_import",
            "mileage",
            "job_name",
            "assigned_to_user_id",
            "assigned_to_email",
            "assigned_to_username",
            "assigned_shop_id",
            "assigned_shop_name",
            "assigned_shop_email",
            "date_done",
            "documents_text",
            "photos_text",
            "note",
            "created_at",
        ],
        [
            [
                str(report.pk),
                str(report.car_id),
                _car_reference(report.car),
                _string_or_empty(report.mileage),
                report.job_name,
                _string_or_empty(report.assigned_to_id),
                report.assigned_to.email if report.assigned_to else "",
                report.assigned_to.username if report.assigned_to else "",
                _string_or_empty(report.assigned_shop_id),
                report.assigned_shop.name if report.assigned_shop else "",
                report.assigned_shop.email if report.assigned_shop else "",
                _format_date(report.date_done),
                _join_list(report.documents),
                _join_list(report.photos),
                report.note,
                _format_datetime(report.created_at),
            ]
            for report in reports
        ],
    )

    content = _serialize_workbook(workbook)
    filename = _build_filename(garage, exported_at)
    return GarageExportWorkbook(filename=filename, content=content)


def _append_rows(sheet: Any, header: list[str], rows: list[list[Any]]) -> None:
    sheet.append(header)
    for row in rows:
        sheet.append(row)


def _serialize_workbook(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _string_or_empty(value: Any) -> str:
    return "" if value in (None, "") else str(value)


def _format_datetime(value: datetime | None) -> str:
    return "" if value is None else value.isoformat()


def _format_date(value: Any) -> str:
    return "" if value is None else value.isoformat()


def _join_list(value: Any) -> str:
    if isinstance(value, list):
        return "\n".join(str(item).strip() for item in value if str(item).strip())
    if value in (None, ""):
        return ""
    return str(value)


def _car_reference(car: Car) -> str:
    return car.vin or car.license_plate or car.usual_name or str(car.pk)


def _build_filename(garage: Garage, exported_at: datetime) -> str:
    slug = "".join(ch.lower() if ch.isalnum() else "-" for ch in garage.name).strip("-")
    compact_slug = "-".join(part for part in slug.split("-") if part) or str(garage.pk)
    timestamp = exported_at.strftime("%Y%m%d_%H%M%S")
    return f"garage_{compact_slug}_{timestamp}.xlsx"
